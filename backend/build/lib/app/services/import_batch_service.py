from __future__ import annotations

import csv
import io
import re
from typing import Any, Optional
from uuid import UUID

from fastapi import HTTPException, UploadFile, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import hash_password
from app.db.models import AcademicCycle, ImportBatch, Student, Tutor, User
from app.db.models.enums import DisciplineEnum, RoleEnum, StudentLifecycleStatusEnum
from app.services.access import can_mutate_lifecycle, ensure_discipline_scope
from app.services.audit_service import record_audit
from app.services.student_service import _student_state_dict


def _normalize_header(h: str) -> str:
    return re.sub(r"\s+", " ", str(h).strip().lower())


def _parse_file(content: bytes, filename: str) -> tuple[list[str], list[dict[str, Any]]]:
    name = filename.lower()
    if name.endswith(".csv"):
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        if reader.fieldnames is None:
            return [], []
        headers = [str(h) for h in reader.fieldnames]
        rows = []
        for row in reader:
            rows.append({str(k): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k})
        return headers, rows
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise HTTPException(status_code=500, detail="openpyxl is required for xlsx") from e

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(c).strip() if c is not None else "" for c in header_row]
    rows: list[dict[str, Any]] = []
    for raw in rows_iter:
        if all(v is None or str(v).strip() == "" for v in raw):
            continue
        row_dict: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = raw[i] if i < len(raw) else None
            row_dict[h] = val
        rows.append(row_dict)
    return headers, rows


def _map_row(
    raw: dict[str, Any],
    mapping: dict[str, str],
) -> dict[str, Any]:
    """Map file headers (keys in mapping) -> canonical field names (values)."""
    out: dict[str, Any] = {}
    for file_key, canon in mapping.items():
        if file_key in raw and raw[file_key] is not None:
            v = raw[file_key]
            if isinstance(v, float) and v == int(v):
                v = int(v)
            out[canon] = str(v).strip() if not isinstance(v, (dict, list)) else v
    return out


def _validate_discipline(val: str) -> bool:
    return val in {e.value for e in DisciplineEnum}


async def process_student_batch(
    session: AsyncSession,
    *,
    actor: User,
    file: UploadFile,
    mapping: dict[str, str],
    dry_run: bool,
    default_password: str,
) -> dict[str, Any]:
    if not can_mutate_lifecycle(actor):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")

    content = await file.read()
    headers, rows = _parse_file(content, file.filename or "upload.csv")
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows in file")

    mapped_rows: list[dict[str, Any]] = []
    for raw in rows:
        mapped_rows.append(_map_row(raw, mapping))

    codes_in_file: list[str] = []
    row_results: list[dict[str, Any]] = []
    for i, data in enumerate(mapped_rows, start=2):
        errs: list[str] = []
        email = data.get("email")
        student_code = data.get("student_code")
        discipline = data.get("discipline", "").strip().lower()
        if not email:
            errs.append("email is required")
        if not student_code:
            errs.append("student_code is required")
        if not discipline:
            errs.append("discipline is required")
        elif not _validate_discipline(discipline):
            errs.append("invalid discipline")
        else:
            try:
                ensure_discipline_scope(actor, discipline)
            except HTTPException:
                errs.append("discipline not allowed for actor")

        ls = data.get("lifecycle_status", StudentLifecycleStatusEnum.pending_onboarding.value)
        if ls not in {e.value for e in StudentLifecycleStatusEnum}:
            errs.append("invalid lifecycle_status")

        if student_code:
            sc = student_code.strip()
            if sc in codes_in_file:
                errs.append("duplicate student_code in file")
            else:
                codes_in_file.append(sc)

        if student_code:
            existing = await session.execute(select(Student).where(Student.student_code == student_code.strip()))
            if existing.scalar_one_or_none():
                errs.append("student_code exists in database")

        existing_e = await session.execute(select(User).where(User.email == email))
        if email and existing_e.scalar_one_or_none():
            errs.append("email exists in database")

        ok = len(errs) == 0
        row_results.append({"row_number": i, "ok": ok, "errors": errs, "data": data})

    batch_record: ImportBatch | None = None
    if dry_run:
        batch_record = ImportBatch(
            batch_type="students",
            file_name=file.filename,
            status="completed",
            details={"dry_run": True, "rows": row_results},
            created_by=actor.id,
        )
        session.add(batch_record)
        await session.commit()
        return {"dry_run": True, "rows": row_results, "import_batch_id": str(batch_record.id)}

    if any(not r["ok"] for r in row_results):
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "rows": row_results})

    batch_record = ImportBatch(
        batch_type="students",
        file_name=file.filename,
        status="processing",
        details={"rows": row_results},
        created_by=actor.id,
    )
    session.add(batch_record)
    await session.flush()

    created_ids: list[str] = []
    for data in mapped_rows:
        discipline = data["discipline"].strip().lower()
        cycle_id: Optional[UUID] = None
        if data.get("academic_cycle_name"):
            r = await session.execute(
                select(AcademicCycle).where(AcademicCycle.name == data["academic_cycle_name"])
            )
            c = r.scalar_one_or_none()
            if c:
                cycle_id = c.id

        user = User(
            email=data["email"],
            hashed_password=hash_password(default_password),
            full_name=data.get("full_name"),
            role=RoleEnum.student.value,
            discipline=discipline,
            is_active=True,
            is_verified=True,
            created_by=actor.id,
        )
        session.add(user)
        await session.flush()

        st = Student(
            user_id=user.id,
            student_code=data["student_code"].strip(),
            institution=data.get("institution"),
            lifecycle_status=data.get("lifecycle_status", StudentLifecycleStatusEnum.pending_onboarding.value),
            discipline=discipline,
            academic_cycle_id=cycle_id,
            created_by=actor.id,
        )
        session.add(st)
        await session.flush()
        created_ids.append(str(st.id))
        await record_audit(
            session,
            actor_id=actor.id,
            action="CREATE",
            entity_type="student",
            entity_id=st.id,
            before_state=None,
            after_state=_student_state_dict(st, user),
            metadata={"batch_import": str(batch_record.id)},
        )

    batch_record.status = "completed"
    batch_record.details = {"created_student_ids": created_ids, "rows": row_results}
    await session.commit()
    return {"dry_run": False, "created_count": len(created_ids), "import_batch_id": str(batch_record.id)}


async def process_tutor_batch(
    session: AsyncSession,
    *,
    actor: User,
    file: UploadFile,
    mapping: dict[str, str],
    dry_run: bool,
    default_password: str,
) -> dict[str, Any]:
    if not can_mutate_lifecycle(actor):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Insufficient permissions")

    content = await file.read()
    _headers, rows = _parse_file(content, file.filename or "upload.csv")
    if not rows:
        raise HTTPException(status_code=400, detail="No data rows in file")

    mapped_rows = [_map_row(raw, mapping) for raw in rows]

    codes_in_file: list[str] = []
    row_results: list[dict[str, Any]] = []
    for i, data in enumerate(mapped_rows, start=2):
        errs: list[str] = []
        email = data.get("email")
        tutor_code = data.get("tutor_code")
        discipline = data.get("discipline", "").strip().lower()
        if not email:
            errs.append("email is required")
        if not tutor_code:
            errs.append("tutor_code is required")
        if not discipline:
            errs.append("discipline is required")
        elif not _validate_discipline(discipline):
            errs.append("invalid discipline")
        else:
            try:
                ensure_discipline_scope(actor, discipline)
            except HTTPException:
                errs.append("discipline not allowed for actor")

        if tutor_code:
            tc = tutor_code.strip()
            if tc in codes_in_file:
                errs.append("duplicate tutor_code in file")
            else:
                codes_in_file.append(tc)

        if tutor_code:
            existing = await session.execute(select(Tutor).where(Tutor.tutor_code == tutor_code.strip()))
            if existing.scalar_one_or_none():
                errs.append("tutor_code exists in database")

        existing_e = await session.execute(select(User).where(User.email == email))
        if email and existing_e.scalar_one_or_none():
            errs.append("email exists in database")

        ok = len(errs) == 0
        row_results.append({"row_number": i, "ok": ok, "errors": errs, "data": data})

    if dry_run:
        batch_record = ImportBatch(
            batch_type="tutors",
            file_name=file.filename,
            status="completed",
            details={"dry_run": True, "rows": row_results},
            created_by=actor.id,
        )
        session.add(batch_record)
        await session.commit()
        return {"dry_run": True, "rows": row_results, "import_batch_id": str(batch_record.id)}

    if any(not r["ok"] for r in row_results):
        raise HTTPException(status_code=400, detail={"message": "Validation failed", "rows": row_results})

    batch_record = ImportBatch(
        batch_type="tutors",
        file_name=file.filename,
        status="processing",
        details={"rows": row_results},
        created_by=actor.id,
    )
    session.add(batch_record)
    await session.flush()

    created_ids: list[str] = []
    for data in mapped_rows:
        discipline = data["discipline"].strip().lower()
        cycle_id: Optional[UUID] = None
        if data.get("academic_cycle_name"):
            r = await session.execute(
                select(AcademicCycle).where(AcademicCycle.name == data["academic_cycle_name"])
            )
            c = r.scalar_one_or_none()
            if c:
                cycle_id = c.id

        user = User(
            email=data["email"],
            hashed_password=hash_password(default_password),
            full_name=data.get("full_name"),
            role=RoleEnum.tutor.value,
            discipline=discipline,
            is_active=True,
            is_verified=True,
            created_by=actor.id,
        )
        session.add(user)
        await session.flush()

        tu = Tutor(
            user_id=user.id,
            tutor_code=data["tutor_code"].strip(),
            discipline=discipline,
            academic_cycle_id=cycle_id,
            created_by=actor.id,
        )
        session.add(tu)
        await session.flush()
        created_ids.append(str(tu.id))
        await record_audit(
            session,
            actor_id=actor.id,
            action="CREATE",
            entity_type="tutor",
            entity_id=tu.id,
            before_state=None,
            after_state={
                "id": str(tu.id),
                "tutor_code": tu.tutor_code,
                "email": user.email,
                "discipline": tu.discipline,
            },
            metadata={"batch_import": str(batch_record.id)},
        )

    batch_record.status = "completed"
    batch_record.details = {"created_tutor_ids": created_ids, "rows": row_results}
    await session.commit()
    return {"dry_run": False, "created_count": len(created_ids), "import_batch_id": str(batch_record.id)}
